# -*- coding: utf-8 -*-
import enum
import os
import json, requests
import hashlib
import xlrd
import xlwt
import openpyxl
import pandas
import time


def main():
    print('-----start translate-----')
    # 获取当前执行文件所在文件夹
    currentFolder = os.path.split(os.path.realpath(__file__))[0] + '/'
    fileName = 'NewChen.xlsx'

    # # 指定列翻译
    # sheetTargetInsert = {
    #     'P&L': [2],
    #     # 'Sheet2': ['小文', 2]
    # }
    # insertTransResult(currentFolder, fileName, sheetTargetInsert)
    
    # 全sheet表翻译
    sheetTargetTotal = {
        'P&L1': ['_all'],
        # 'Sheet2': ['A', 'C'],
        # 'P&L1': ['C']
    }
    totalTransResult(currentFolder, fileName, sheetTargetTotal)
    
# 读取excel内容,在指定文件、指定sheet的指定列后新增一列，存放翻译结果
# currentFolder: 当前文件夹
# fileName: 文件名
# sheetTarget: 需要翻译的sheet及对应列数据
def insertTransResult(currentFolder, fileName, sheetTarget):
    # 获取所有sheet表
    sourceFile = currentFolder + fileName
    pdSheetList = pandas.read_excel(sourceFile, sheet_name=None)

    pdWriter = pandas.ExcelWriter(currentFolder + 'trans_' + fileName)
    excelBook = xlrd.open_workbook(sourceFile, encoding_override="utf-8")

    for sheetName in list(pdSheetList):
        # 如果sheet需要翻译，则进行翻译和生成新excel
        if(sheetName in list( sheetTarget.keys() )):
            # 获取表的索引值
            sheetIndex = list(pdSheetList).index(sheetName)
            # 获取指定的excel sheet表
            sheetData = excelBook.sheets()[sheetIndex]

            # 获取指定sheet表的数据
            pdSheet = pandas.read_excel(sourceFile, sheetName)
            colNames = pdSheet.columns.tolist()
            
            flag = 0
            for column in sheetTarget[sheetName]:
                print('正在翻译列:', column)
                # 根据column类型进行处理
                if isinstance(column, str):
                    colIndex = colNames.index(column)
                    colName = column
                else:
                    colIndex = column
                    colName = ''
                
                # 翻译列名
                transColName = trans_words('it', 'zh', colName)
                transResultList = []
                
                # 循环遍历进行翻译
                for transWords in sheetData.col(colIndex)[1:]:
                    time.sleep(0.1)
                    transResult = trans_words('it', 'zh', transWords.value)
                    print('翻译结果:',transWords.value, '--', transResult)
                    transResultList.append(transResult)

                # 寻找指定列插入数据
                pdSheet.insert(colIndex + flag + 1, transColName + ('un' if transColName=='' else '(trans)'), transResultList, allow_duplicates=True)
                flag = flag + 1
            pdSheet.to_excel(pdWriter, sheet_name=sheetName, index=False)
            print(sheetName+' sheet已生成')

    pdWriter.save()
    print('trans_' + fileName + '表已生成完毕')
    return


# 读取excel内容,在指定文件、指定sheet表全文翻译
# currentFolder: 当前文件夹
# fileName: 文件名
# sheetName: sheet表名称
def totalTransResult(currentFolder, fileName, sheetTarget):
    sourceFile = currentFolder + fileName

    excelBook = openpyxl.load_workbook(sourceFile)
    for sheetName in sheetTarget.keys():
        print('当前处理表:',sheetName)
        try:
            sheet = excelBook[sheetName]
        except Exception as e:
            print(e)
            continue
        # 数组仅有_all且数组长度为1时即全表翻译
        if(len(sheetTarget[sheetName]) == 1 and sheetTarget[sheetName][0] == '_all'):
            for col in range(sheet.max_column):
                for row in range(sheet.max_row):
                    transWords = sheet.cell(row+1, col+1).value 
                    if (transWords != '' and transWords is not None):
                        transResult = trans_words('it', 'zh', transWords)
                        print('翻译结果:', transWords, '--', transResult )
                        transResult = '' if transResult == '' else '(' + transResult + ')'
                        sheet.cell(row+1, col+1, transWords + transResult)
        # 翻译指定列的内容
        else:
            for col in sheetTarget[sheetName]:
                for row in range(sheet.max_row):
                    transWords = sheet[col+str(row+1)].value
                    if (transWords != '' and transWords is not None):
                        transResult = trans_words('it', 'zh', transWords)
                        print('翻译结果:', transWords, '--', transResult )
                        transResult = '' if transResult == '' else '(' + transResult + ')'
                        sheet[col+str(row+1)] = transWords + transResult
 
    excelBook.save(currentFolder+'translate_'+fileName)
    print("表格新增完成")


# 翻译字符
# 参数 fromText：源语言， toText：目标语言，transWords： 待翻译内容
def trans_words(fromText, toText, transWords):
    # http://api.fanyi.baidu.com/api/trans/vip/translate?q=banana&from=en&to=zh&appid=20220331001153008&salt=1435660288&sign=125eedcc4b2d7db300c11503710fe10b
    if(transWords == '' or  transWords is None):
        return ""

    transUrl = 'http://api.fanyi.baidu.com/api/trans/vip/translate'
    salt = '1435660288' 
    password = 'xxxxx'
    appid = '123213212321' 

    sign = hashlib.md5((appid+transWords+salt+password).encode(encoding='UTF-8')).hexdigest()

    # 请求参数拼接
    data = {
        "from": fromText,   # 源语言 en-英语 kor-韩文 jp-日语 it-意大利 zh 中文
        "to": toText,       # 翻译目标语言
        "q": transWords,    # 待翻译字符
        "appid": appid,     # 开发者id
        "salt": salt,       # 随机数
        "sign": sign,       # 签名：appid+q+salt+密钥的顺序拼接
    }

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36',
    }
    try:
        res = requests.post(transUrl, data=data, headers=headers, timeout=3).json()
        # print(res)
        data_result = json.loads(json.dumps(res, indent=2, ensure_ascii=False)) # dumps将字符串转化为json， loads将json转化为字典
    except Exception as e:
        print('请求出现异常:', e)
        return ""


    if('trans_result' not in data_result.keys()):
        return ""
    return data_result['trans_result'][0]['dst']

if __name__ =='__main__':
    main()
